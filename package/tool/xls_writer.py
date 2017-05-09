#!/usr/bin/env python

from openpyxl import Workbook
import json,os
import datetime

class XLSWriter:
    def __init__(self, ):
        pass;
    def __del__(self,):
        pass;
    def write(self, resultCollection,resultFolder):
        self.resultFolder = resultFolder;
        fname = resultFolder + str(datetime.datetime.now().strftime("%Y%m%d%HH%MM")) +".xlsx";
        workbook = Workbook();
        worksheet = workbook.worksheets[0]
        worksheet.title = "Sheet1"
        worksheet.cell(row=0, column=0).value = "URL";
        worksheet.cell(row=0, column=1).value = "AVLIST";
        worksheet.cell(row=0, column=2).value = "SCORE";
        worksheet.cell(row=0, column=3).value = "MODULE";
        worksheet.cell(row=0, column=4).value = "STATUS";
        column = 0;
        rowi =1;
        for key, val in resultCollection.items():
            jsontxtval = json.loads(val);
            avlist = jsontxtval['avList'];
            score = jsontxtval['score'];
            module = jsontxtval['module'];
            status = jsontxtval['status'];
            worksheet.cell(row=rowi, column=0).value= key.encode('utf-8');
            worksheet.cell(row=rowi, column=1).value= str(avlist);
            worksheet.cell(row=rowi, column=2).value= str(score);
            worksheet.cell(row=rowi, column=3).value= str(module);
            worksheet.cell(row=rowi, column=4).value= str(status);
            rowi +=1;
        workbook.save(fname);
        return fname;

